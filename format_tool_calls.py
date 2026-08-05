"""Plot pie charts of tool call distribution from Harbor trial runs.

Definitions (read from ``trial/agent/trajectory.json``):

- **Step**: one entry in ``trajectory["steps"]`` — a single agent turn
  (LLM response plus its tool execution results). Identified by ``step_id``.
  Steps that contain no tool calls are ignored.
- **Tool call**: one entry in ``step["tool_calls"]`` — a single function
  invocation requested by the model (e.g. one ``bash_command``). A single
  step can contain multiple tool calls when the model emits them in
  parallel, so tool-calls-per-trial ≥ steps-per-trial.

Usage:
    cd examples/otel-demo
    uv run python format_tool_calls.py -jd jobs-0218
    uv run python format_tool_calls.py -jd jobs-0218 -od out
"""

import json
import re
from collections import Counter
from pathlib import Path

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from matplotlib.colors import to_hex
from openpyxl.styles import PatternFill
from tabulate import tabulate

from harbor_utils.plotting_utils import (
    LABEL_FONT_SIZE,
    TICK_FONT_SIZE,
)
from utils import (
    CODE_CATEGORIES,
    TELEM_ALL_CATEGORIES,
    TELEM_DATA_CATEGORIES,
    classify_cmd,
    find_trial_dirs,
    format_group_label,
    get_base_parser,
    load_invalid_trial_ids,
)

# ── Shell tool helpers ────────────────────────────────────────────────
SHELL_TOOL_NAMES = {"exec_command", "bash_command"}


def _extract_shell_cmd(function_name: str, arguments: dict) -> str:
    """Return the shell command string for a shell-style tool call, else ''."""
    if function_name == "exec_command":
        return arguments.get("cmd", "")
    if function_name == "bash_command":
        # terminus-2 sends raw keystrokes incl. trailing newline
        return arguments.get("keystrokes", "").rstrip("\n")
    return ""


# ── Language detection ─────────────────────────────────────────────────
LANG_REGEXES = [
    ("YAML", r"\.yaml\b|\.yml\b"),
    ("Go", r"\.go\b"),
    ("TypeScript", r"\.ts\b"),
    ("JavaScript", r"\.js\b"),
    ("Python", r"\.py\b"),
    ("Java", r"\.java\b"),
    ("C#", r"\.cs\b"),
    ("Ruby", r"\.rb\b"),
    ("PHP", r"\.php\b"),
    ("Rust", r"\.rs\b"),
    ("Protobuf", r"\.proto\b"),
    ("Docker", r"Dockerfile"),
    ("Kotlin", r"\.kt\b"),
]


def _extract_tool_calls_from_trajectory(trajectory: dict) -> list[dict]:
    """Extract tool call rows from a single trajectory dict."""
    rows: list[dict] = []
    for step in trajectory.get("steps", []):
        tool_calls = step.get("tool_calls", [])
        if not tool_calls:
            continue

        # Build a lookup from source_call_id -> content
        results = step.get("observation", {}).get("results", [])
        output_by_id: dict[str, str] = {}
        shared_output: str | None = None
        for r in results:
            if "source_call_id" in r:
                output_by_id[r["source_call_id"]] = r.get("content", "")
            else:
                shared_output = r.get("content", "")

        for tc in tool_calls:
            tc_id = tc.get("tool_call_id", "")
            output = output_by_id.get(tc_id, shared_output or "")
            function_name = tc.get("function_name", "unknown")
            arguments = tc.get("arguments", {})
            categories = (
                classify_cmd(_extract_shell_cmd(function_name, arguments))
                if function_name in SHELL_TOOL_NAMES
                else None
            )
            rows.append(
                {
                    "step_id": step.get("step_id", ""),
                    "tool_call_id": tc_id,
                    "function_name": function_name,
                    "arguments": json.dumps(arguments),
                    "output": output,
                    "categories": categories,
                }
            )
    return rows


def collect_tool_calls(
    jobs_dir: Path,
    csv_dir: Path | None = None,
    valid_trial_ids: set[str] | None = None,
) -> list[dict]:
    """Extract individual tool calls from all trajectory files.

    Args:
        jobs_dir: Path to the Harbor jobs directory.
        csv_dir: If provided, save a separate CSV per trajectory to this dir.
        valid_trial_ids: If provided, skip any trajectory whose trial_id is
            not in this set (neither rows collected nor CSV written).

    Returns:
        List of dicts with keys: group, trial_id, step_id, tool_call_id,
        function_name, arguments, output.

    """
    jobs_dir = jobs_dir.resolve()
    if not jobs_dir.exists():
        raise FileNotFoundError(f"Jobs directory not found: {jobs_dir}")

    if csv_dir is not None:
        csv_dir.mkdir(parents=True, exist_ok=True)

    all_rows: list[dict] = []

    for batch_dir in sorted(jobs_dir.iterdir()):
        if not batch_dir.is_dir():
            continue

        agent_name, model_name = "unknown", "unknown"
        effort: str | None = None
        config_path = batch_dir / "config.json"
        if config_path.exists():
            try:
                config = json.loads(config_path.read_text())
                if config.get("agents"):
                    agent_cfg = config["agents"][0]
                    agent_name = agent_cfg.get("name", "unknown")
                    model_name = agent_cfg.get("model_name", "unknown")
                    effort = (agent_cfg.get("kwargs") or {}).get("reasoning_effort")
            except Exception:
                pass

        group = format_group_label(agent_name, model_name, effort)

        for trial_dir in sorted(batch_dir.iterdir()):
            if not trial_dir.is_dir():
                continue

            trial_id = trial_dir.name
            if valid_trial_ids is not None and trial_id not in valid_trial_ids:
                continue

            trajectory_path = trial_dir / "agent" / "trajectory.json"
            if not trajectory_path.exists():
                continue

            try:
                with open(trajectory_path) as f:
                    trajectory = json.load(f)
            except Exception as e:
                print(f"Skipping {trial_dir}: {e}")
                continue

            rows = _extract_tool_calls_from_trajectory(trajectory)

            if csv_dir is not None and rows:
                pd.DataFrame(rows).to_csv(csv_dir / f"{trial_id}.csv", index=False)

            for row in rows:
                row["group"] = group
                row["trial_id"] = trial_id
            all_rows.extend(rows)

    return all_rows


def analyze_run(run_name: str, df: pd.DataFrame) -> dict:
    """Analyze a single run and return a dict of computed fields."""
    exec_rows = df[df["function_name"].isin(SHELL_TOOL_NAMES)]

    cmds = []
    for _, row in exec_rows.iterrows():
        try:
            args = json.loads(row["arguments"])
            cmds.append(_extract_shell_cmd(row["function_name"], args))
        except Exception:
            cmds.append("")

    # Classify each command
    cmd_categories = [classify_cmd(c) for c in cmds]
    flat_categories = [c for cats in cmd_categories for c in cats]

    # Language detection
    langs = set()
    for cmd in cmds:
        for lang_name, pattern in LANG_REGEXES:
            if re.search(pattern, cmd):
                langs.add(lang_name)

    # Parse run name (strip __hash suffix from CSV filenames)
    clean_name = re.sub(r"__\w+$", "", run_name)
    m = re.match(r"(\w+(?:-\w+)*?)-(\d+)-(lag\w+)-(avail\w+)", clean_name)
    task_type = m.group(1) if m else ""
    task_idx = m.group(2) if m else ""
    lag = m.group(3) if m else ""
    avail = m.group(4) if m else ""

    # First telemetry data signal
    first_telem_data = None
    for cats in cmd_categories:
        for c in cats:
            if c in TELEM_DATA_CATEGORIES:
                first_telem_data = c
                break
        if first_telem_data:
            break

    # Starts with code or telemetry
    starts_with = None
    for cats in cmd_categories:
        is_code = any(c in CODE_CATEGORIES for c in cats)
        is_telem = any(c in TELEM_ALL_CATEGORIES for c in cats)
        if is_code and not is_telem:
            starts_with = "code"
            break
        elif is_telem and not is_code:
            starts_with = "telemetry"
            break
        elif is_code and is_telem:
            starts_with = "both"
            break

    # High-level sequence and transitions
    high_seq = []
    for cats in cmd_categories:
        is_code = any(c in CODE_CATEGORIES for c in cats)
        is_telem = any(c in TELEM_ALL_CATEGORIES for c in cats)
        label = (
            "both"
            if (is_code and is_telem)
            else "code"
            if is_code
            else "telemetry"
            if is_telem
            else "other"
        )
        if not high_seq or high_seq[-1] != label:
            high_seq.append(label)

    transitions = sum(
        1
        for i in range(1, len(high_seq))
        if (high_seq[i - 1] == "code" and high_seq[i] == "telemetry")
        or (high_seq[i - 1] == "telemetry" and high_seq[i] == "code")
    )

    # Code/telemetry percentages
    total_cmds = len(cmd_categories)
    n_code = sum(
        1 for cats in cmd_categories if any(c in CODE_CATEGORIES for c in cats)
    )
    n_telem = sum(
        1 for cats in cmd_categories if any(c in TELEM_ALL_CATEGORIES for c in cats)
    )

    # Quartile code density
    quartile_code = []
    for qi in range(4):
        ss = total_cmds / 4
        start = int(qi * ss)
        end = int((qi + 1) * ss)
        chunk = cmd_categories[start:end]
        code_count = sum(1 for cats in chunk if any(c in CODE_CATEGORIES for c in cats))
        quartile_code.append(code_count / max(len(chunk), 1))

    # Error count
    n_errors = 0
    for _, row in exec_rows.iterrows():
        output = str(row.get("output", ""))
        if (
            "Process exited with code 1" in output
            or "Process exited with code 2" in output
        ):
            n_errors += 1

    return {
        "run_name": run_name,
        "task_type": task_type,
        "task_idx": task_idx,
        "lag": lag,
        "avail": avail,
        "total_calls": len(df),
        "all_cmds": cmds,
        "cmd_categories": cmd_categories,
        "flat_categories": flat_categories,
        "languages": langs,
        "n_languages": len(langs),
        "first_telem_data": first_telem_data,
        "starts_with": starts_with,
        "high_level_seq": high_seq,
        "code_telem_transitions": transitions,
        "n_code_cmds": n_code,
        "n_telem_cmds": n_telem,
        "pct_code": n_code / max(total_cmds, 1),
        "pct_telem": n_telem / max(total_cmds, 1),
        "quartile_code_density": quartile_code,
        "n_errors": n_errors,
        "error_rate": n_errors / max(len(cmds), 1),
    }


def main() -> None:
    """Plot pie charts of tool call distribution per agent."""
    parser = get_base_parser()
    parser.description = "Plot pie charts of tool call distribution per agent."
    parser.add_argument(
        "--xlsx",
        action="store_true",
        help="Also write per-group tool_calls-*.xlsx files (skipped by default).",
    )
    args = parser.parse_args()

    csv_dir = args.output_dir / "tool_calls"
    invalid_ids: set[str] = (
        load_invalid_trial_ids(args.filter_xlsx) if args.filter_xlsx else set()
    )

    # Restrict to trials the analysis layer considers scored — the same set
    # utils.load_trials builds — so tool-call stats cover exactly the trials
    # that appear in the accuracy tables.
    valid_trial_ids = {
        d.name
        for d in find_trial_dirs(args.jobs_dir)
        if (d / "verifier" / "reward-details.json").is_file()
    }
    print(f"Restricting to {len(valid_trial_ids)} scored trial(s) in {args.jobs_dir}")

    rows = collect_tool_calls(
        args.jobs_dir, csv_dir=csv_dir, valid_trial_ids=valid_trial_ids
    )
    if invalid_ids:
        rows = [r for r in rows if r["trial_id"] not in invalid_ids]

    if not rows:
        print("No tool calls found.")
        return
    print(f"Saved per-trajectory CSVs to {csv_dir}/")

    df = pd.DataFrame(rows)
    exec_mask = df["function_name"].isin(SHELL_TOOL_NAMES)
    exec_df = df[exec_mask].copy()
    groups = sorted(df["group"].unique())

    # Stable category -> color map, shared between Excel cells and pie charts.
    # Warm tones for code:*, cool tones for telemetry:*, grey for the rest.
    # Within each prefix, the most-frequent category gets the most-saturated shade.
    global_cat_counter: Counter[str] = Counter()
    for cats in exec_df["categories"]:
        global_cat_counter.update(cats)
    ordered_cats = [c for c, _ in global_cat_counter.most_common()]

    _ramps = {
        "code": ("YlOrRd", 0.9, 0.35),
        "telemetry": ("GnBu", 0.9, 0.35),
    }
    _by_prefix: dict[str, list[str]] = {}
    for cat in ordered_cats:
        _by_prefix.setdefault(cat.split(":")[0], []).append(cat)

    cat_colors: dict[str, tuple] = {}
    for prefix, cats in _by_prefix.items():
        name, lo, hi = _ramps.get(prefix, ("Greys", 0.7, 0.3))
        ramp = plt.get_cmap(name)
        n = len(cats)
        for i, cat in enumerate(cats):
            pos = lo if n <= 1 else lo + (hi - lo) * i / (n - 1)
            cat_colors[cat] = ramp(pos)
    cat_argb = {
        cat: "FF" + to_hex(c, keep_alpha=False)[1:].upper()
        for cat, c in cat_colors.items()
    }

    # Save one Excel file per group, with one sheet per trajectory, row-colored by primary category.
    _ansi_re = re.compile(r"\x1b\[[0-9;]*m")
    _illegal_xml_re = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]")
    _hash_re = re.compile(r"(__\w+)$")

    def _safe_name(s: str) -> str:
        return re.sub(r"[^A-Za-z0-9._-]+", "_", s)

    def _sheet_name(trial_id: str) -> str:
        """Truncate to Excel's 31-char limit while preserving the __hash suffix."""
        s = str(trial_id)
        m = _hash_re.search(s)
        if not m:
            return s[:31]
        suffix = m.group(1)
        prefix = s[: m.start()]
        return (prefix[: max(31 - len(suffix), 0)] + suffix)[:31]

    args.output_dir.mkdir(parents=True, exist_ok=True)
    if args.xlsx:
        for group in groups:
            group_df = df[df["group"] == group]
            if group_df.empty:
                continue
            xlsx_path = args.output_dir / f"tool_calls-{_safe_name(group)}.xlsx"
            with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
                for trial_id, trial_df in group_df.groupby("trial_id"):
                    sheet = _sheet_name(trial_id)
                    clean = trial_df.drop(columns=["group", "trial_id"]).copy()
                    # Strip ANSI escapes and illegal XML characters (openpyxl rejects them)
                    clean["output"] = clean["output"].apply(
                        lambda s: _illegal_xml_re.sub("", _ansi_re.sub("", s))
                    )
                    clean.to_excel(writer, sheet_name=sheet, index=False)
                    ws = writer.sheets[sheet]
                    n_cols = clean.shape[1]
                    for i, cats in enumerate(clean["categories"]):
                        if not cats:
                            continue
                        argb = cat_argb.get(cats[0])
                        if argb is None:
                            continue
                        fill = PatternFill(
                            start_color=argb, end_color=argb, fill_type="solid"
                        )
                        for col in range(1, n_cols + 1):
                            ws.cell(row=i + 2, column=col).fill = fill
            print(f"Saved Excel to {xlsx_path}")

    # Print category overlap table per group
    def _print_overlap(label: str, sub_df: pd.DataFrame) -> None:
        total = len(sub_df)
        counter: Counter[str] = Counter()
        for cats in sub_df["categories"]:
            prefixes = sorted(set(c.split(":")[0] for c in cats))
            if len(prefixes) > 1:
                counter[" + ".join(prefixes)] += 1
        multi = sum(counter.values())
        rows = [
            [combo, count, f"{100 * count / max(total, 1):.1f}%"]
            for combo, count in counter.most_common()
        ]
        rows.append(["**Total**", multi, f"{100 * multi / max(total, 1):.1f}%"])
        print(f"\n[{label}] Multi-category overlaps ({multi}/{total} shell cmds):")
        print(
            tabulate(
                rows,
                headers=["Combo", "Count", "% of exec"],
                tablefmt="github",
            )
        )
        print()

    for group in groups:
        _print_overlap(group, exec_df[exec_df["group"] == group])

    # Plot pie charts — function_name distribution
    n_groups = len(groups)

    fig, axes = plt.subplots(1, n_groups, figsize=(4 * n_groups, 4), squeeze=False)

    for i, group in enumerate(groups):
        ax = axes[0, i]
        counter = Counter(df[df["group"] == group]["function_name"])
        labels, sizes = zip(*counter.most_common())
        ax.pie(
            sizes,
            labels=labels,
            autopct="%1.1f%%",
            textprops={"fontsize": TICK_FONT_SIZE},
        )
        ax.set_title(group, fontsize=LABEL_FONT_SIZE)

    fig.suptitle("Tool Call Distribution per Agent", fontsize=LABEL_FONT_SIZE)
    fig.tight_layout()

    fig_path = args.output_dir / "figures" / "tool_calls_distribution.pdf"
    fig_path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(fig_path, bbox_inches="tight")
    print(f"Saved figure to {fig_path}")

    # Histogram — number of tool calls per trial, overlaid per group.
    counts_df = (
        df.groupby(["group", "trial_id"]).size().reset_index(name="n_tool_calls")
    )
    fig_h, ax_h = plt.subplots(figsize=(5, 3))
    bins = 20
    lo = counts_df["n_tool_calls"].min()
    hi = counts_df["n_tool_calls"].max()
    bin_edges = [lo + (hi - lo) * i / bins for i in range(bins + 1)]
    for group in groups:
        subset = counts_df[counts_df["group"] == group]["n_tool_calls"]
        if subset.empty:
            continue
        patches = ax_h.hist(
            subset,
            bins=bin_edges,
            label=group,
            alpha=0.5,
            edgecolor="white",
            linewidth=0.5,
        )
        color = patches[2][0].get_facecolor()
        ax_h.axvline(subset.median(), color=color, linestyle="--", linewidth=1.2)
    ax_h.set_xlabel("Tool calls per trial", fontsize=LABEL_FONT_SIZE)
    ax_h.set_ylabel("Count", fontsize=LABEL_FONT_SIZE)
    ax_h.tick_params(labelsize=TICK_FONT_SIZE)
    ax_h.legend(fontsize=TICK_FONT_SIZE)
    fig_h.suptitle("Distribution of Tool Calls per Trial", fontsize=LABEL_FONT_SIZE)
    fig_h.tight_layout()
    hist_path = args.output_dir / "figures" / "tool_calls_count_histogram.pdf"
    fig_h.savefig(hist_path, bbox_inches="tight")
    print(f"Saved figure to {hist_path}")

    # Histogram — number of steps per trial, one row per group.
    steps_df = (
        df.groupby(["group", "trial_id"])["step_id"]
        .nunique()
        .reset_index(name="n_steps")
    )
    nonempty_step_groups = [
        g for g in groups if not steps_df[steps_df["group"] == g].empty
    ]
    n_rows = len(nonempty_step_groups)
    fig_s, axes_s = plt.subplots(
        n_rows, 1, figsize=(5, 2 * n_rows), sharex=True, squeeze=False
    )
    lo = steps_df["n_steps"].min()
    hi = steps_df["n_steps"].max()
    bin_edges = [lo + (hi - lo) * i / bins for i in range(bins + 1)]
    for ax_s, group in zip(axes_s[:, 0], nonempty_step_groups):
        subset = steps_df[steps_df["group"] == group]["n_steps"]
        ax_s.hist(
            subset,
            bins=bin_edges,
            label=group,
            alpha=0.7,
            edgecolor="white",
            linewidth=0.5,
        )
        ax_s.axvline(subset.median(), color="black", linestyle="--", linewidth=1.2)
        ax_s.set_ylabel("Count", fontsize=LABEL_FONT_SIZE)
        ax_s.tick_params(labelsize=TICK_FONT_SIZE)
        ax_s.legend(fontsize=TICK_FONT_SIZE)
    axes_s[-1, 0].set_xlabel("Steps per trial", fontsize=LABEL_FONT_SIZE)
    fig_s.suptitle("Distribution of Steps per Trial", fontsize=LABEL_FONT_SIZE)
    fig_s.tight_layout()
    steps_path = args.output_dir / "figures" / "steps_count_histogram.pdf"
    fig_s.savefig(steps_path, bbox_inches="tight")
    print(f"Saved figure to {steps_path}")

    # Plot pie charts — category distribution for exec_command calls.
    fig2, axes2 = plt.subplots(1, n_groups, figsize=(5 * n_groups, 5), squeeze=False)

    for i, group in enumerate(groups):
        ax = axes2[0, i]
        group_cats = exec_df[exec_df["group"] == group]["categories"]
        counter = Counter(cat for cats in group_cats for cat in cats)
        if counter:
            labels = [c for c in ordered_cats if c in counter]
            sizes = [counter[c] for c in labels]
            colors = [cat_colors[la] for la in labels]
            ax.pie(
                sizes,
                labels=labels,
                colors=colors,
                autopct="%1.1f%%",
                textprops={"fontsize": TICK_FONT_SIZE},
            )
        ax.set_title(group, fontsize=LABEL_FONT_SIZE)

    fig2.suptitle(
        "exec_command Category Distribution per Agent", fontsize=LABEL_FONT_SIZE
    )
    fig2.tight_layout()

    fig2_path = args.output_dir / "figures" / "tool_calls_category_distribution.pdf"
    fig2.savefig(fig2_path, bbox_inches="tight")
    print(f"Saved category figure to {fig2_path}")

    # Print category summary per group
    for group in groups:
        sub = exec_df[exec_df["group"] == group]
        cats_counter: Counter[str] = Counter()
        for cats in sub["categories"]:
            cats_counter.update(cats)
        total_assignments = sum(cats_counter.values())
        total_g = len(sub)
        print(f"\n[{group}] Total shell cmds: {total_g}")
        print(
            f"[{group}] Total category assignments: {total_assignments}"
            f" ({total_assignments / max(total_g, 1):.2f} per cmd)"
        )
        print(f"[{group}] Category distribution:")
        for cat, count in cats_counter.most_common():
            pct = 100 * count / max(total_assignments, 1)
            print(f"  {cat}: {count} ({pct:.1f}%)")

    # Per-run analysis — group runs by agent-model
    csv_dir = args.output_dir / "tool_calls"
    trial_to_group = dict(zip(df["trial_id"], df["group"]))
    runs_by_group: dict[str, list[dict]] = {g: [] for g in groups}
    for csv_path in sorted(csv_dir.glob("*.csv")):
        trial_id = csv_path.stem
        if args.filter_xlsx and trial_id in invalid_ids:
            continue
        group = trial_to_group.get(trial_id)
        if group is None:
            continue
        run_df = pd.read_csv(csv_path)
        runs_by_group[group].append(analyze_run(trial_id, run_df))

    nonempty_groups = [g for g in groups if runs_by_group[g]]
    for group in nonempty_groups:
        all_runs = runs_by_group[group]
        n_runs = len(all_runs)
        print(f"\n[{group}] Per-run analysis ({n_runs} runs):")
        print(
            f"  Avg languages per run: "
            f"{np.mean([r['n_languages'] for r in all_runs]):.2f}"
        )
        print(
            f"  Avg transitions per run: "
            f"{np.mean([r['code_telem_transitions'] for r in all_runs]):.2f}"
        )
        print(
            f"  Avg error rate: "
            f"{100 * np.mean([r['error_rate'] for r in all_runs]):.1f}%"
        )

    # Workflow start patterns: share of trials whose first command is
    # code-only, plus the fine-grained category of the first telemetry-data
    # command. Denominator for the per-category columns is trials that
    # issued at least one telemetry-data command; columns appear only for
    # categories that show up in the data, ordered canonically.
    def _cell(n: int, total: int) -> str:
        return f"{100 * n / total:.1f}% ({n}/{total})" if total else "—"

    telem_cats_present: set[str] = set()
    for group in nonempty_groups:
        for r in runs_by_group[group]:
            if r["first_telem_data"] is not None:
                telem_cats_present.add(r["first_telem_data"])
    canonical_order = [
        "telemetry:metrics",
        "telemetry:logs",
        "telemetry:docker_logs",
        "telemetry:spans",
        "telemetry:grafana_query_unresolved",
    ]
    telem_cats = [c for c in canonical_order if c in telem_cats_present] + sorted(
        telem_cats_present - set(canonical_order)
    )

    start_rows = []
    for group in nonempty_groups:
        all_runs = runs_by_group[group]
        n_runs = len(all_runs)
        n_start_code = sum(1 for r in all_runs if r["starts_with"] == "code")
        runs_with_telem = [r for r in all_runs if r["first_telem_data"] is not None]
        n_telem = len(runs_with_telem)
        first_cat_counts = Counter(r["first_telem_data"] for r in runs_with_telem)
        start_rows.append(
            [
                group,
                n_runs,
                _cell(n_start_code, n_runs),
            ]
            + [_cell(first_cat_counts[c], n_telem) for c in telem_cats]
        )
    print(
        "\nWorkflow start patterns "
        "(Starts-with-code over all trials; First telem = X over trials with any telemetry-data call):"
    )
    print(
        tabulate(
            start_rows,
            headers=["Agent", "Trials", "Starts with code"]
            + [f"First telem = {c.split(':', 1)[1]}" for c in telem_cats],
            tablefmt="github",
        )
    )

    # Fine-grained code-subcategory breakdown for the trials that start with
    # code. A single command can carry multiple ``code:*`` labels, so each
    # cell counts multi-label commands once per label and rows can sum to
    # slightly more than 100%. Denominator is the number of starts-with-code
    # trials in that group.
    def _first_code_categories(cmd_categories: list[list[str]]) -> list[str]:
        """Return the ``code:*`` labels of the command that set ``starts_with``."""
        for cats in cmd_categories:
            is_code = any(c in CODE_CATEGORIES for c in cats)
            is_telem = any(c in TELEM_ALL_CATEGORIES for c in cats)
            if is_code and not is_telem:
                return [c for c in cats if c in CODE_CATEGORIES]
            if is_telem:
                return []
        return []

    code_subcats = sorted(CODE_CATEGORIES)
    code_rows = []
    for group in nonempty_groups:
        code_starts = [r for r in runs_by_group[group] if r["starts_with"] == "code"]
        if not code_starts:
            continue
        n = len(code_starts)
        counter: Counter[str] = Counter()
        for r in code_starts:
            counter.update(_first_code_categories(r["cmd_categories"]))
        code_rows.append([group, n] + [_cell(counter[sc], n) for sc in code_subcats])
    if code_rows:
        print("\nFirst code-command categories (among starts-with-code trials):")
        print(
            tabulate(
                code_rows,
                headers=["Agent", "n"] + code_subcats,
                tablefmt="github",
            )
        )

    # Combined language figure: grouped bars per agent-model, normalized to
    # density (% of runs within each group).
    if nonempty_groups:
        per_group_lang_counts = {
            g: [r["n_languages"] for r in runs_by_group[g]] for g in nonempty_groups
        }
        per_group_lang_freq: dict[str, Counter[str]] = {}
        for g in nonempty_groups:
            c: Counter[str] = Counter()
            for r in runs_by_group[g]:
                c.update(r["languages"])
            per_group_lang_freq[g] = c

        max_lang = max(max(v) for v in per_group_lang_counts.values())
        x_ticks = np.arange(0, max_lang + 1)
        bins = np.arange(-0.5, max_lang + 1.5, 1)

        # Union of languages across groups, ordered by total frequency.
        total_lang_freq: Counter[str] = Counter()
        for c in per_group_lang_freq.values():
            total_lang_freq.update(c)
        all_languages = [la for la, _ in total_lang_freq.most_common()]
        x_pos = np.arange(len(all_languages))

        n_g = len(nonempty_groups)
        cmap = plt.get_cmap("tab10")
        group_colors = {g: cmap(i % 10) for i, g in enumerate(nonempty_groups)}

        fig3, (ax_hist, ax_bar) = plt.subplots(1, 2, figsize=(13, 4.5))

        # Panel (a): grouped bars over languages-touched-per-run
        bar_w_a = 0.8 / n_g
        for i, g in enumerate(nonempty_groups):
            counts = per_group_lang_counts[g]
            n = len(counts)
            hist, _ = np.histogram(counts, bins=bins)
            pcts = 100 * hist / n
            offset = (i - (n_g - 1) / 2) * bar_w_a
            ax_hist.bar(
                x_ticks + offset,
                pcts,
                width=bar_w_a,
                color=group_colors[g],
                edgecolor="white",
                label=f"{g} (n={n})",
            )
            ax_hist.axvline(
                np.median(counts),
                color=group_colors[g],
                linestyle="--",
                linewidth=1.2,
            )
        ax_hist.set_title("(a) Languages touched per run", fontsize=10)
        ax_hist.set_xlabel("Number of languages touched", fontsize=9)
        ax_hist.set_ylabel("% of runs", fontsize=9)
        ax_hist.set_xticks(x_ticks)
        ax_hist.set_xlim(-0.6, max_lang + 0.6)
        ax_hist.spines["top"].set_visible(False)
        ax_hist.spines["right"].set_visible(False)
        ax_hist.tick_params(labelsize=8)
        ax_hist.legend(fontsize=7)

        # Panel (b): grouped bars over per-language frequency
        bar_w_b = 0.8 / n_g
        for i, g in enumerate(nonempty_groups):
            n = len(runs_by_group[g])
            freq = per_group_lang_freq[g]
            pcts = [100 * freq.get(la, 0) / n for la in all_languages]
            offset = (i - (n_g - 1) / 2) * bar_w_b
            ax_bar.bar(
                x_pos + offset,
                pcts,
                width=bar_w_b,
                color=group_colors[g],
                edgecolor="white",
                label=f"{g} (n={n})",
            )
        ax_bar.set_title("(b) Language frequency across runs", fontsize=10)
        ax_bar.set_xlabel("Language", fontsize=9)
        ax_bar.set_ylabel("% of runs touching language", fontsize=9)
        ax_bar.set_xticks(x_pos)
        ax_bar.set_xticklabels(all_languages, rotation=40, ha="right", fontsize=8)
        ax_bar.spines["top"].set_visible(False)
        ax_bar.spines["right"].set_visible(False)
        ax_bar.tick_params(labelsize=8)
        ax_bar.legend(fontsize=7)

        fig3.tight_layout()
        fig3_path = args.output_dir / "figures" / "languages.pdf"
        fig3.savefig(fig3_path, dpi=200, bbox_inches="tight")
        print(f"Saved language figure to {fig3_path}")

    plt.show()


if __name__ == "__main__":
    main()
